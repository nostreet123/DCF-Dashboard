from __future__ import annotations

import base64
import csv
import importlib
import io
import re
import uuid
import zipfile
from dataclasses import dataclass
from datetime import date
from typing import Any

import xlrd
from openpyxl import load_workbook

from dcf_engine.service.company_contracts import (
    ImportParseRequest,
    ImportParseResponse,
    ImportReview,
    ImportReviewField,
    ImportedArtifactKind,
    ImportedArtifactMetadata,
    ParseArtifactRequest,
    ParsedFieldCandidate,
    ParsedFieldName,
)


REQUIRED_FIELDS: tuple[ParsedFieldName, ...] = (
    "periodEnd",
    "filingCurrency",
    "revenue",
    "cash",
    "debt",
    "sharesOutstanding",
)

MAX_TABULAR_ROWS = 2_000
MAX_TABULAR_COLUMNS = 64
MAX_CELL_TEXT_LENGTH = 10_000
MAX_SPREADSHEET_SHEETS = 8
MAX_SPREADSHEET_ZIP_MEMBERS = 256
MAX_SPREADSHEET_UNCOMPRESSED_BYTES = 24 * 1024 * 1024
MAX_PDF_PAGES = 40
MAX_PDF_TEXT_CHARS = 200_000


FIELD_LABELS: dict[ParsedFieldName, tuple[str, ...]] = {
    "periodEnd": ("period end", "year end", "fiscal year", "date"),
    "filingDate": ("filing date", "publication date", "report date"),
    "filingCurrency": ("currency", "reporting currency", "presentation currency"),
    "revenue": ("revenue", "sales", "turnover", "net sales"),
    "cash": ("cash and cash equivalents", "cash equivalents", "cash balance", "cash at bank"),
    "debt": ("total debt", "borrowings", "loans payable", "financial debt", "debt balance"),
    "sharesOutstanding": ("shares outstanding", "ordinary shares", "number of shares"),
}


@dataclass(frozen=True)
class CellRow:
    cells: list[str]

    @property
    def text(self) -> str:
        return " ".join(cell for cell in self.cells if cell).strip()


def infer_kind(filename: str, preferred: ImportedArtifactKind | None) -> ImportedArtifactKind:
    if preferred is not None:
        return preferred
    normalized = filename.casefold()
    if "income" in normalized or "profit" in normalized or "p&l" in normalized:
        return "incomeStatement"
    if "balance" in normalized or "position" in normalized:
        return "balanceSheet"
    if "cash" in normalized and "flow" in normalized:
        return "cashFlow"
    if "share" in normalized or "currency" in normalized or "meta" in normalized:
        return "sharesMeta"
    return "incomeStatement"


def parse_number(raw: str) -> float | None:
    cleaned = raw.strip()
    if not cleaned:
        return None
    negative = cleaned.startswith("(") and cleaned.endswith(")")
    cleaned = cleaned.strip("()")
    multiplier = 1.0
    if re.search(r"\b(million|mn|m)\b", cleaned, flags=re.IGNORECASE):
        multiplier = 1_000_000.0
    elif re.search(r"\b(thousand|k)\b", cleaned, flags=re.IGNORECASE):
        multiplier = 1_000.0
    numeric = re.sub(r"[^0-9.,\-]", "", cleaned)
    if not numeric:
        return None
    if numeric.count(",") > 0 and numeric.count(".") == 0:
        numeric = numeric.replace(",", ".") if numeric.count(",") == 1 else numeric.replace(",", "")
    else:
        numeric = numeric.replace(",", "")
    try:
        value = float(numeric) * multiplier
    except ValueError:
        return None
    return -value if negative else value


def normalize_date(raw: str) -> str | None:
    match = re.search(r"(20\d{2}|19\d{2})[-/.](1[0-2]|0?[1-9])[-/.](3[01]|[12]\d|0?[1-9])", raw)
    if match:
        year, month, day = match.groups()
        return f"{int(year):04d}-{int(month):02d}-{int(day):02d}"
    match = re.search(r"(3[01]|[12]\d|0?[1-9])[-/.](1[0-2]|0?[1-9])[-/.](20\d{2}|19\d{2})", raw)
    if match:
        day, month, year = match.groups()
        return f"{int(year):04d}-{int(month):02d}-{int(day):02d}"
    match = re.search(r"\b(20\d{2}|19\d{2})\b", raw)
    if match:
        return f"{int(match.group(1)):04d}-12-31"
    return None


def period_for_value_cell(
    cells: list[str],
    value_index: int | None,
    period_by_column: dict[int, str],
    fallback: str | None,
) -> str | None:
    if value_index is None:
        return fallback
    cell_period = normalize_date(cells[value_index])
    if cell_period:
        return cell_period
    if value_index in period_by_column:
        return period_by_column[value_index]
    for cell in reversed(cells[:value_index]):
        cell_period = normalize_date(cell)
        if cell_period:
            return cell_period
    return fallback


def _trim_cell(value: Any) -> str:
    text = "" if value is None else str(value).strip()
    return text[:MAX_CELL_TEXT_LENGTH]


def _limited_csv_rows(content: bytes, delimiter: str) -> list[CellRow]:
    reader = csv.reader(io.StringIO(content.decode("utf-8-sig", errors="replace")), delimiter=delimiter)
    rows: list[CellRow] = []
    for row_number, row in enumerate(reader, start=1):
        if row_number > MAX_TABULAR_ROWS:
            raise ValueError(f"Import file has too many rows; maximum {MAX_TABULAR_ROWS}")
        if len(row) > MAX_TABULAR_COLUMNS:
            raise ValueError(f"Import file has too many columns; maximum {MAX_TABULAR_COLUMNS}")
        rows.append(CellRow([_trim_cell(cell) for cell in row]))
    return rows


def _validate_xlsx_zip(content: bytes) -> None:
    try:
        with zipfile.ZipFile(io.BytesIO(content)) as workbook_zip:
            infos = workbook_zip.infolist()
    except zipfile.BadZipFile as exc:
        raise ValueError("Invalid XLSX import file") from exc
    if len(infos) > MAX_SPREADSHEET_ZIP_MEMBERS:
        raise ValueError(f"XLSX import has too many archive entries; maximum {MAX_SPREADSHEET_ZIP_MEMBERS}")
    uncompressed_bytes = sum(info.file_size for info in infos)
    if uncompressed_bytes > MAX_SPREADSHEET_UNCOMPRESSED_BYTES:
        raise ValueError("XLSX import expands beyond the supported size limit")


def _read_xlsx_rows(content: bytes) -> list[CellRow]:
    _validate_xlsx_zip(content)
    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    try:
        sheet_names = workbook.sheetnames
        if len(sheet_names) > MAX_SPREADSHEET_SHEETS:
            raise ValueError(f"Spreadsheet import has too many sheets; maximum {MAX_SPREADSHEET_SHEETS}")
        rows: list[CellRow] = []
        for sheet_name in sheet_names:
            worksheet = workbook[sheet_name]
            if worksheet.max_row is not None and worksheet.max_row > MAX_TABULAR_ROWS:
                raise ValueError(f"Spreadsheet import has too many rows; maximum {MAX_TABULAR_ROWS}")
            if worksheet.max_column is not None and worksheet.max_column > MAX_TABULAR_COLUMNS:
                raise ValueError(f"Spreadsheet import has too many columns; maximum {MAX_TABULAR_COLUMNS}")
            rows.append(CellRow([str(sheet_name)]))
            for row_number, row in enumerate(
                worksheet.iter_rows(max_col=MAX_TABULAR_COLUMNS, values_only=True),
                start=1,
            ):
                if row_number > MAX_TABULAR_ROWS:
                    raise ValueError(f"Spreadsheet import has too many rows; maximum {MAX_TABULAR_ROWS}")
                rows.append(CellRow([_trim_cell(cell) for cell in row]))
        return rows
    finally:
        workbook.close()


def _read_xls_rows(content: bytes) -> list[CellRow]:
    workbook = xlrd.open_workbook(file_contents=content, on_demand=True)
    try:
        if workbook.nsheets > MAX_SPREADSHEET_SHEETS:
            raise ValueError(f"Spreadsheet import has too many sheets; maximum {MAX_SPREADSHEET_SHEETS}")
        rows: list[CellRow] = []
        for sheet_name in workbook.sheet_names():
            worksheet = workbook.sheet_by_name(sheet_name)
            if worksheet.nrows > MAX_TABULAR_ROWS:
                raise ValueError(f"Spreadsheet import has too many rows; maximum {MAX_TABULAR_ROWS}")
            if worksheet.ncols > MAX_TABULAR_COLUMNS:
                raise ValueError(f"Spreadsheet import has too many columns; maximum {MAX_TABULAR_COLUMNS}")
            rows.append(CellRow([str(sheet_name)]))
            for row_index in range(worksheet.nrows):
                rows.append(
                    CellRow(
                        [
                            _trim_cell(worksheet.cell_value(row_index, column_index))
                            for column_index in range(worksheet.ncols)
                        ]
                    )
                )
        return rows
    finally:
        workbook.release_resources()


def read_tabular_rows(content: bytes, filename: str) -> list[CellRow]:
    suffix = filename.rsplit(".", 1)[-1].casefold() if "." in filename else ""
    if suffix in {"csv", "tsv"}:
        sample = content[:4096].decode("utf-8-sig", errors="replace")
        if suffix == "tsv":
            delimiter = "\t"
        else:
            try:
                delimiter = csv.Sniffer().sniff(sample, delimiters=",;\t").delimiter
            except csv.Error:
                delimiter = ","
        return _limited_csv_rows(content, delimiter)
    if suffix == "xlsx":
        return _read_xlsx_rows(content)
    if suffix == "xls":
        return _read_xls_rows(content)
    return []


def _load_pdf_reader() -> type[Any]:
    return importlib.import_module("pypdf").PdfReader


def read_pdf_text(content: bytes) -> str:
    reader = _load_pdf_reader()(io.BytesIO(content))
    if len(reader.pages) > MAX_PDF_PAGES:
        raise ValueError(f"PDF import has too many pages; maximum {MAX_PDF_PAGES}")
    text_parts: list[str] = []
    text_chars = 0
    for page in reader.pages:
        page_text = page.extract_text() or ""
        text_chars += len(page_text)
        if text_chars > MAX_PDF_TEXT_CHARS:
            raise ValueError("PDF import extracted too much text")
        text_parts.append(page_text)
    return "\n".join(text_parts)


def candidates_from_rows(
    rows: list[CellRow],
    *,
    artifact_id: str,
    filename: str,
    review_required: bool,
) -> list[ParsedFieldCandidate]:
    candidates: list[ParsedFieldCandidate] = []
    latest_period = None
    period_by_column: dict[int, str] = {}
    for row in rows:
        row_text = row.text
        if not row_text:
            continue
        row_lower = row_text.casefold()
        row_period = normalize_date(row_text)
        for index, cell in enumerate(row.cells):
            cell_period = normalize_date(cell)
            if cell_period:
                period_by_column[index] = cell_period
        latest_period = max(filter(None, [latest_period, row_period]), default=None)
        for field, labels in FIELD_LABELS.items():
            if not any(label in row_lower for label in labels):
                continue
            value = None
            numeric = None
            value_index = None
            if field in {"periodEnd", "filingDate"}:
                value = normalize_date(row_text)
            elif field == "filingCurrency":
                currency = re.search(r"\b[A-Z]{3}\b", row_text)
                value = currency.group(0) if currency else None
            else:
                for index in range(len(row.cells) - 1, -1, -1):
                    cell = row.cells[index]
                    numeric = parse_number(cell)
                    if numeric is not None:
                        value = cell
                        value_index = index
                        break
            if not value:
                continue
            candidate_period = period_for_value_cell(
                row.cells,
                value_index,
                period_by_column,
                row_period or latest_period,
            )
            candidates.append(
                ParsedFieldCandidate(
                    field=field,
                    rawValue=value,
                    numericValue=numeric,
                    periodEnd=candidate_period,
                    artifactId=artifact_id,
                    artifactFilename=filename,
                    confidence="medium" if review_required else "high",
                    reviewRequired=review_required,
                )
            )
    return candidates


def candidates_from_text(
    text: str,
    *,
    artifact_id: str,
    filename: str,
    review_required: bool,
) -> list[ParsedFieldCandidate]:
    rows = [CellRow([line.strip()]) for line in text.splitlines()]
    return candidates_from_rows(
        rows,
        artifact_id=artifact_id,
        filename=filename,
        review_required=review_required,
    )


def parse_artifact(request: ParseArtifactRequest) -> ImportedArtifactMetadata:
    artifact_id = request.id or str(uuid.uuid4())
    content = base64.b64decode(request.content_base64)
    filename = request.original_filename
    suffix = filename.rsplit(".", 1)[-1].casefold() if "." in filename else ""
    kind = infer_kind(filename, request.preferred_kind)
    notes: list[str] = []
    review_required = suffix == "pdf"

    if suffix in {"csv", "tsv", "xlsx", "xls"}:
        rows = read_tabular_rows(content, filename)
        candidates = candidates_from_rows(
            rows,
            artifact_id=artifact_id,
            filename=filename,
            review_required=review_required,
        )
        parser_name = "Excel" if suffix in {"xlsx", "xls"} else "CSV"
    elif suffix == "pdf":
        text = read_pdf_text(content)
        candidates = candidates_from_text(
            text,
            artifact_id=artifact_id,
            filename=filename,
            review_required=True,
        )
        parser_name = "PDF"
        notes.append("PDF-derived fields require manual confirmation before approval.")
    else:
        raise ValueError(f"Unsupported import format: {suffix or 'unknown'}")

    if not candidates:
        notes.append("No recognizable valuation fields were detected.")

    return ImportedArtifactMetadata(
        id=artifact_id,
        kind=kind,
        originalFilename=filename,
        parserName=parser_name,
        fileFormat=suffix,
        requiresReview=review_required,
        candidates=candidates,
        notes=notes,
    )


def _candidate_sort_key(candidate: ParsedFieldCandidate) -> tuple[str, int]:
    confidence_score = {"high": 3, "medium": 2, "low": 1}[candidate.confidence]
    return (candidate.period_end or "", confidence_score)


def build_review(artifacts: list[ImportedArtifactMetadata]) -> ImportReview:
    candidates = [candidate for artifact in artifacts for candidate in artifact.candidates]
    chosen_period = max(
        (candidate.period_end for candidate in candidates if candidate.period_end),
        default=None,
    )
    fields: list[ImportReviewField] = []
    for field in FIELD_LABELS:
        matches = [candidate for candidate in candidates if candidate.field == field]
        if chosen_period:
            period_matches = [
                candidate
                for candidate in matches
                if candidate.period_end == chosen_period or field in {"periodEnd", "filingDate", "filingCurrency"}
            ]
            if period_matches:
                matches = period_matches
        if not matches:
            continue
        candidate = sorted(matches, key=_candidate_sort_key, reverse=True)[0]
        fields.append(
            ImportReviewField(
                field=field,
                value=candidate.raw_value,
                sourceFilename=candidate.artifact_filename,
                confidence=candidate.confidence,
                isManualOverride=False,
                confirmed=not candidate.review_required,
            )
        )

    field_map = {field.field: field.value for field in fields}
    missing = [
        field
        for field in REQUIRED_FIELDS
        if not field_map.get(field)
        or (field in {"revenue", "sharesOutstanding"} and (parse_number(field_map[field]) or 0) <= 0)
        or (field in {"cash", "debt"} and parse_number(field_map[field]) is None)
        or (field == "periodEnd" and normalize_date(field_map[field]) is None)
    ]
    notes = sorted({note for artifact in artifacts for note in artifact.notes})
    return ImportReview(
        chosenPeriodEnd=chosen_period,
        fields=fields,
        missingRequiredFields=missing,
        notes=notes,
        isValuationReady=not missing and all(
            field.confirmed or field.field not in REQUIRED_FIELDS for field in fields
        ),
    )


def parse_import_payload(payload: ImportParseRequest) -> ImportParseResponse:
    artifacts = [parse_artifact(artifact) for artifact in payload.artifacts]
    review = build_review(artifacts)
    return ImportParseResponse(
        artifacts=artifacts,
        candidates=[candidate for artifact in artifacts for candidate in artifact.candidates],
        review=review,
    )


def make_imported_statement_from_review(review: dict[str, Any]) -> dict[str, Any]:
    field_values = {
        field.get("field"): str(field.get("value", "")).strip()
        for field in review.get("fields", [])
        if isinstance(field, dict)
    }
    period_end = normalize_date(field_values.get("periodEnd", "")) or date.today().isoformat()
    return {
        "periodEnd": period_end,
        "periodType": "FY",
        "filingDate": normalize_date(field_values.get("filingDate", "")),
        "currency": field_values.get("filingCurrency") or None,
        "revenue": parse_number(field_values.get("revenue", "")),
        "cash": parse_number(field_values.get("cash", "")),
        "debt": parse_number(field_values.get("debt", "")),
        "sharesOutstanding": parse_number(field_values.get("sharesOutstanding", "")),
        "source": "import",
    }
